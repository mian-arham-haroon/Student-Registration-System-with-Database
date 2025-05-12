from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox 
from PIL import Image , ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"
#------------------------------------Gender-------------------------------------------#

def selection ():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"
    print(gender)
#------------------------------------Exit window-------------------------------------------#
def Exit ():
    root.destroy()
#---------------------------------show image----------------------------------------------#
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetypes=(("jpg file","*.jpg"),
    ("png file","*.png"),
    ("all file","*.txt")))

    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
#------------------------------------Registration No-------------------------------------------#
# now each time we have to enter Registration no.
#let designe automate Registration entry system
# it is created to automatic entry Registration System no

def registration_on():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell (row=row,column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")
#-------------------------------------Clear----------------------------------------------#
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_name.set('')
    M_name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")

    registration_on()
    saveButton.config(state='normal')

    img1=PhotoImage(file="images/upload photo.png")
    lbl.config(image=img1)
    lbl.image=img1
    img=""
#-----------------------------------save--------------------------------------------#
def Save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender") 

    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    Fathername=F_name.get()
    Mothername=M_name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or Fathername=="" or Mothername=="" or F1=="" or M1=="":
        messagebox.showerror("error","Few Data Is Missing")

    else:
        file=openpyxl.load_workbook('Student_data.xlsx') 
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)   
        sheet.cell(column=3,row=sheet.max_row,value=C1)   
        sheet.cell(column=4,row=sheet.max_row,value=G1)   
        sheet.cell(column=6,row=sheet.max_row,value=D2)   
        sheet.cell(column=5,row=sheet.max_row,value=D1)   
        sheet.cell(column=7,row=sheet.max_row,value=Re1)   
        sheet.cell(column=8,row=sheet.max_row,value=S1)   
        sheet.cell(column=9,row=sheet.max_row,value=Fathername)   
        sheet.cell(column=10,row=sheet.max_row,value=Mothername)   
        sheet.cell(column=11,row=sheet.max_row,value=F1)   
        sheet.cell(column=12,row=sheet.max_row,value=M1)   

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student Images/"+str(R1)+".jpg")

        except:
            messagebox.showinfo("info","Perfile Picture is not available!!!")

        messagebox.showinfo("info","Successfully data entered!!!")
        Clear()
        registration_on()
#------------------------------------Search-------------------------------------------#
def search():
    text=Search.get()       #taking input from entry box
    Clear()         #to clear all data
    saveButton.config(state='disabled')  #after clicking on search , save button will disable so that  one click on it

    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            # print(str(name))
            reg_no_position=str(name)[14:-1]        #call sheet a2
            reg_number=str(name)[15:-1]                    #call sheet 2
            # print(reg_no_position)
            # print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number!!!")

    # reg_no_position showing like a2,a3,a4...an
    #but reg_number just showing number after a2 like 2,3,.......n
     
    x1=sheet.cell(row=int(reg_number),column=1).value                 
    x2=sheet.cell(row=int(reg_number),column=2).value                 
    x3=sheet.cell(row=int(reg_number),column=3).value                 
    x4=sheet.cell(row=int(reg_number),column=4).value                 
    x5=sheet.cell(row=int(reg_number),column=5).value                 
    x6=sheet.cell(row=int(reg_number),column=6).value                 
    x7=sheet.cell(row=int(reg_number),column=7).value                 
    x8=sheet.cell(row=int(reg_number),column=8).value                 
    x9=sheet.cell(row=int(reg_number),column=9).value                 
    x10=sheet.cell(row=int(reg_number),column=10).value                 
    x11=sheet.cell(row=int(reg_number),column=11).value                 
    x12=sheet.cell(row=int(reg_number),column=12).value                 

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4=='Female':
        R2.select()
    else:
        R1.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_name.set(x9)
    M_name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img=(Image.open("Student Images/"+str(x1)+".jpg"))  #we will done image name ,save as registration no. 
    resize_image=img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resize_image)
    lbl.config(image=photo2)
    lbl.image=photo2
#------------------------------------Update-------------------------------------------#
def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    Fathername=F_name.get()
    Mothername=M_name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()  

    file=openpyxl.load_workbook('Student_data.xlsx')  
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(Name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
            print(reg_number)

    # sheet.cell(column=1,row=int(reg_number),value=R1)     #now noon can update registration .,it will                                        
    sheet.cell(column=2,row=int(reg_number),value=N1)                                            
    sheet.cell(column=3,row=int(reg_number),value=C1)                                            
    sheet.cell(column=4,row=int(reg_number),value=G1)                                            
    sheet.cell(column=5,row=int(reg_number),value=D2)                                            
    sheet.cell(column=6,row=int(reg_number),value=D1)                                            
    sheet.cell(column=7,row=int(reg_number),value=Re1)                                            
    sheet.cell(column=8,row=int(reg_number),value=S1)                                            
    sheet.cell(column=9,row=int(reg_number),value=Fathername)                                            
    sheet.cell(column=10,row=int(reg_number),value=Mothername)                                            
    sheet.cell(column=11,row=int(reg_number),value=F1)                                            
    sheet.cell(column=12,row=int(reg_number),value=M1)                                            
    file.save(r'Student_data.xlsx')

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update","Update Successfully!!")
    Clear()    

#-------------------------------------------------------------------------------#
root=Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)
#--------------------------------create and save xlsxfile-----------------------------------------------#
file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date Of Registration"
    sheet['G1']="Religion"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father Occupation"
    sheet['L1']="Mother Occupation"

    file.save('Student_data.xlsx')

#-------------------------------------top frame------------------------------------------#
Label(root,text="Email: arhamharoon000@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="Student Registration",width=10,height=2,bg="#c36464",fg='#fff',font='arial 20 bold').pack(side=TOP,fill=X)
#-------------------------------------search box to update------------------------------------------#
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=870,y=70)
image_icon3=PhotoImage(file="images/search.png")
Scrh=Button(root,text="Search",compound=LEFT,image=image_icon3,width=123
            ,bg="#68ddfa",font="arial 13 bold",command=search)
Scrh.place(x=1060,y=66)

image_icon4=PhotoImage(file="images/Layer 4.png")
Update_button=Button(root,image=image_icon4,bg="#c36464",command=Update)
Update_button.place(x=110,y=64)
#--------------------------------------Registration and Data-----------------------------------------#
Label(root,text="Registration No:",font="arial 13",fg=framebg,bg=framefg).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=framefg).place(x=500,y=150)
Registration=IntVar()
Date=StringVar()
reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)
registration_on()
#----------------------------------registration no---------------------------------------------#
today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)
Date.set(d1)
#-------------------------------------Student details------------------------------------------#
# Create the first LabelFrame for "Student's Details"
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50) 
Label(obj,text="Date Of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100) 
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150) 

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50) 
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100) 
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150) 

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion_entry.place(x=630,y=100)

Skill=StringVar()
skill_entry=Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=630,y=150)

Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="roboto 10",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")

radio= IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)

#-------------------------------------Parent details------------------------------------------#
# Create the second LabelFrame for "Parent's Details"
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

F_name=StringVar()
F_entry=Entry(obj2,textvariable=F_name,width=20,font="arial 10")
F_entry.place(x=160,y=50)

Father_Occupation=StringVar()
Fo_entry=Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
Fo_entry.place(x=160,y=100)

Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

M_name=StringVar()
M_entry=Entry(obj2,textvariable=M_name,width=20,font="arial 10")
M_entry.place(x=630,y=50)

Mother_Occupation=StringVar()
Mo_entry=Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
Mo_entry.place(x=630,y=100)
#-----------------------------image--------------------------------------------------#
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)
#---------------------------------Button----------------------------------------------#
Button(root,text="Update",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)
#-------------------------------------------------------------------------------#








#-------------------------------------------------------------------------------#
root.mainloop()